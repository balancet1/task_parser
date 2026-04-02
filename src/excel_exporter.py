import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os
from io import BytesIO

class ExcelExporter:
    def __init__(self, filename=None):
        self.filename = filename
        self.base_filename = "output/tasks.xlsx"
        
        if filename and os.path.exists(filename):
            from openpyxl import load_workbook
            self.wb = load_workbook(filename)
        else:
            self.wb = Workbook()
            if "Sheet" in self.wb.sheetnames:
                self.wb.remove(self.wb["Sheet"])
    
    def add_sheet(self, df, sheet_name: str):
        if sheet_name in self.wb.sheetnames:
            self.wb.remove(self.wb[sheet_name])
        
        ws = self.wb.create_sheet(title=sheet_name)
        
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
                if isinstance(value, (datetime, pd.Timestamp)):
                    cell.number_format = 'DD.MM.YYYY'
        
        self._apply_formatting(ws, len(df.columns), len(df))
    
    def _apply_formatting(self, ws, num_columns, num_rows):
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        date_alignment = Alignment(horizontal='center', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        for row in range(2, num_rows + 2):
            for col in range(1, num_columns + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                
                col_letter = ws.cell(row=1, column=col).value
                if col_letter in ['Срок', 'Дата']:
                    cell.alignment = date_alignment
                else:
                    cell.alignment = cell_alignment
        
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        ws.freeze_panes = 'A2'
    
    def save(self, filename=None):
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"output/tasks_{timestamp}.xlsx"
        
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        
        for sheet in self.wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        _ = cell.value
        
        self.wb.save(filename)
        print(f"✅ Excel файл сохранен: {filename}")
        return filename
    
    def save_to_buffer(self, buffer):
        self.wb.save(buffer)
        return buffer